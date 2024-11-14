import time
import base64
import os
import re
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

def sanitize_filename(filename):
    """Sanitize the filename to remove any characters that are not allowed in file names."""
    return re.sub(r'[\/:*?"<>|]', '', filename)

def clean_linkedin_url(input_url):
    """
    Cleans and validates a LinkedIn job URL.
    Extracts the 10-digit job ID and constructs the correct URL format.
    """
    # Check if the input contains the required LinkedIn jobs string
    if "www.linkedin.com/jobs" not in input_url:
        print("Error: The provided URL is not a valid LinkedIn job URL.")
        return None

    # Use a regular expression to find the 10-digit job ID
    match = re.search(r'(\d{10})', input_url)
    
    if match:
        job_id = match.group(1)
        # Construct the cleaned URL
        cleaned_url = f"https://www.linkedin.com/jobs/view/{job_id}"
        print(f"Cleaned URL: {cleaned_url}")
        return cleaned_url
    else:
        print("Error: Could not find a valid 10-digit job ID in the URL.")
        return None

def check_duplicate(spreadsheet_path, company_name, job_title):
    """Check if a combination of company name and job title already exists in the spreadsheet."""
    if not os.path.exists(spreadsheet_path):
        return False

    workbook = openpyxl.load_workbook(spreadsheet_path)
    sheet = workbook.active

    # Iterate through the rows to check for duplicates
    for row in sheet.iter_rows(min_row=2, values_only=True):
        existing_company, existing_job = row[2], row[3]
        if existing_company == company_name and existing_job == job_title:
            return True

    return False

def update_spreadsheet(date, time, company_name, job_title, file_path):
    """Update the Excel spreadsheet with the job details."""
    spreadsheet_path = r"C:\Users\johns\Documents\Python\Linkedin_bot\Job_Listings.xlsx"

    # Create a new workbook if the spreadsheet doesn't exist
    if not os.path.exists(spreadsheet_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Time", "Company Name", "Job Title", "PDF Link"])
        workbook.save(spreadsheet_path)

    # Check for duplicates
    is_duplicate = check_duplicate(spreadsheet_path, company_name, job_title)
    if is_duplicate:
        user_input = input(f"Duplicate found for '{company_name} - {job_title}'. Do you want to add it anyway? (y/n): ")
        if user_input.lower() != 'y':
            print("Skipped adding the duplicate entry.")
            return

    # Load the existing workbook and add a new row
    workbook = openpyxl.load_workbook(spreadsheet_path)
    sheet = workbook.active
    pdf_link = f'=HYPERLINK("{file_path}", "Open PDF")'
    sheet.append([date, time, company_name, job_title, pdf_link])
    workbook.save(spreadsheet_path)
    print(f"Updated spreadsheet: {spreadsheet_path}")

def save_webpage_as_pdf(url):
    # Set up Selenium WebDriver with Chrome
    options = webdriver.ChromeOptions()
    options.add_argument("user-data-dir=C:/Users/johns/AppData/Local/Google/Chrome/User Data")
    options.add_argument("profile-directory=Default")
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    
    # Specify the save directory
    save_dir = r"C:\Users\johns\Documents\Python\Linkedin_bot"
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.implicitly_wait(5)

    try:
        print(f"Fetching the URL: {url}")
        driver.get(url)

        # Extract the job title
        try:
            job_title_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "h1.t-24.t-bold.inline"))
            )
            job_title = job_title_element.text
            print(f"Job Title: {job_title}")
        except Exception as e:
            print("Failed to get job title:", e)
            job_title = "Unknown Job Title"

        # Extract the company name
        try:
            company_name_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.job-details-jobs-unified-top-card__company-name a.app-aware-link"))
            )
            company_name = company_name_element.text.strip()
            print(f"Company Name: {company_name}")
        except Exception as e:
            print("Failed to get company name:", e)
            company_name = "Unknown Company"

        # Sanitize the job title and company name for the file name
        job_title = sanitize_filename(job_title)
        company_name = sanitize_filename(company_name)

        # Add a timestamp to the filename
        timestamp = datetime.now()
        date_str = timestamp.strftime('%Y-%m-%d')
        time_str = timestamp.strftime('%H-%M')
        output_filename = f"{date_str} -- {company_name} -- {job_title}.pdf"
        output_path = os.path.join(save_dir, output_filename)

        # Click the "See more" button to expand the job description
        try:
            see_more_button = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(@aria-label, 'see more description')]"))
            )
            driver.execute_script("arguments[0].click();", see_more_button)
            print("Clicked the 'See more' button to expand the job description.")
        except Exception as e:
            print("Failed to click 'See more' button:", e)

        # Use Chrome's built-in print-to-PDF feature
        settings = {"printBackground": True}
        result = driver.execute_cdp_cmd("Page.printToPDF", settings)
        
        # Decode the base64 PDF data
        pdf_data = base64.b64decode(result['data'])
        
        # Write the decoded data to a PDF file
        with open(output_path, 'wb') as file:
            file.write(pdf_data)
        print(f"Saved webpage as {output_path}")

        # Update the spreadsheet with the job details
        update_spreadsheet(date_str, time_str, company_name, job_title, output_path)

    finally:
        driver.quit()

if __name__ == "__main__":
    input_url = input("Enter the LinkedIn job URL: ")
    cleaned_url = clean_linkedin_url(input_url)
    
    if cleaned_url:
        save_webpage_as_pdf(cleaned_url)

















